import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------- PARÂMETROS DO CENÁRIO PÉ NO CHÃO ----------------
INVESTIMENTO_INICIAL = 53710      # R$
PRECO_HORA_ESTACAO = 35           # R$/h
HORAS_DIA = 4
DIAS_MES = 20                     # menos dias efetivos
ESTACOES = 4
TAXA_OCUPACAO = 0.60              # 60% de ocupação média

# Custos fixos mensais (mais conservadores)
CUSTOS_FIXOS_MES = {
    'Aluguel': 1633,
    'Condomínio/IPTU': 1000,
    'Internet': 100,
    'Energia': 900,
    'Manutenção': 700,
    'Limpeza/Segurança': 400
}
CUSTOS_FIXOS_TOTAL = sum(CUSTOS_FIXOS_MES.values())

# Receita e lucro mensais no cenário pé no chão
RECEITA_MES_BRUTA = PRECO_HORA_ESTACAO * HORAS_DIA * DIAS_MES * ESTACOES
RECEITA_MES_EFETIVA = RECEITA_MES_BRUTA * TAXA_OCUPACAO
LUCRO_OPERACIONAL_MES = RECEITA_MES_EFETIVA - CUSTOS_FIXOS_TOTAL

TAXA_DESCONTO_ANUAL = 0.12
ANOS_ANALISE = 5

print("=== CENÁRIO PÉ NO CHÃO ===")
print(f"Investimento inicial: R$ {INVESTIMENTO_INICIAL:,.2f}")
print(f"Receita mensal bruta (100% ocupação): R$ {RECEITA_MES_BRUTA:,.2f}")
print(f"Receita mensal efetiva (60%): R$ {RECEITA_MES_EFETIVA:,.2f}")
print(f"Custos fixos mensais: R$ {CUSTOS_FIXOS_TOTAL:,.2f}")
print(f"Lucro operacional mensal: R$ {LUCRO_OPERACIONAL_MES:,.2f}")

# ---------------- FUNÇÕES FINANCEIRAS ----------------
def calcular_vpl(fluxo, taxa):
    return sum(fc / (1 + taxa) ** t for t, fc in enumerate(fluxo))

def calcular_tir(fluxo):
    low, high = -0.99, 10.0
    for _ in range(100):
        mid = (low + high) / 2
        npv = calcular_vpl(fluxo, mid)
        if npv > 0:
            low = mid
        else:
            high = mid
    return low

def calcular_payback(fluxo):
    acumulado = 0
    for i in range(len(fluxo)):
        acumulado += fluxo[i]
        if acumulado >= 0:
            acumulado_antes = acumulado - fluxo[i]
            restante = -acumulado_antes
            fracao = restante / fluxo[i]
            return (i - 1) + fracao
    return None

# ---------------- FLUXO DE CAIXA ANUAL ----------------
anos = list(range(0, ANOS_ANALISE + 1))
fluxo_caixa = [-INVESTIMENTO_INICIAL]
for ano in range(1, ANOS_ANALISE + 1):
    lucro_anual = LUCRO_OPERACIONAL_MES * 12 * (1 - 0.05 * (ano - 1))
    fluxo_caixa.append(lucro_anual)

print("Fluxo de caixa anual (pé no chão):", fluxo_caixa)

# Indicadores
payback_anos = calcular_payback(fluxo_caixa)
vpl_12 = calcular_vpl(fluxo_caixa, TAXA_DESCONTO_ANUAL)
tir = calcular_tir(fluxo_caixa)

print(f"Payback: {payback_anos if payback_anos is not None else 'não recupera em 5 anos'} anos")
print(f"VPL 12%: R$ {vpl_12:,.2f}")
print(f"TIR: {tir*100:.2f}%")

# ---------------- CRIAÇÃO DA PLANILHA ----------------
wb = Workbook()

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
bold_font = Font(bold=True)

# ===== Aba PARÂMETROS =====
ws_p = wb.active
ws_p.title = "Param_Pe_no_Chao"

ws_p["A1"] = "PARÂMETROS - CENÁRIO PÉ NO CHÃO"
ws_p["A1"].font = Font(bold=True, size=14)
ws_p.merge_cells("A1:B1")

ws_p["A2"] = "Descrição"
ws_p["B2"] = "Valor"
for col in ("A", "B"):
    c = ws_p[f"{col}2"]
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

parametros = [
    ("Investimento Inicial (R$)", INVESTIMENTO_INICIAL),
    ("Preço/Hora/Estação (R$)", PRECO_HORA_ESTACAO),
    ("Horas por dia", HORAS_DIA),
    ("Dias por mês", DIAS_MES),
    ("Número de estações", ESTACOES),
    ("Taxa de ocupação", TAXA_OCUPACAO),
    ("Receita mensal bruta (R$)", RECEITA_MES_BRUTA),
    ("Receita mensal efetiva (R$)", RECEITA_MES_EFETIVA),
    ("Custos fixos mensais (R$)", CUSTOS_FIXOS_TOTAL),
    ("Lucro operacional mensal (R$)", LUCRO_OPERACIONAL_MES),
    ("Taxa de desconto anual", TAXA_DESCONTO_ANUAL),
    ("Período de análise (anos)", ANOS_ANALISE),
]

for i, (nome, val) in enumerate(parametros, start=3):
    ws_p[f"A{i}"] = nome
    ws_p[f"A{i}"].font = bold_font
    ws_p[f"B{i}"] = float(val)
    if "R$" in nome:
        ws_p[f"B{i}"].number_format = 'R$ #,##0.00'
    elif "Taxa" in nome:
        ws_p[f"B{i}"].number_format = '0,00%'

ws_p.column_dimensions["A"].width = 40
ws_p.column_dimensions["B"].width = 20

# ===== Aba FLUXO DE CAIXA =====
ws_f = wb.create_sheet("Fluxo_Pe_no_Chao")

headers = ["Ano", "Fluxo de Caixa (R$)", "Fluxo Acumulado (R$)", "Valor Presente (R$)"]
for col, h in enumerate(headers, start=1):
    c = ws_f.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

fluxo_acumulado = 0
for i, ano in enumerate(anos, start=2):
    fc = fluxo_caixa[i-2]
    fluxo_acumulado += fc
    vp = fc / (1 + TAXA_DESCONTO_ANUAL) ** (ano-1)

    ws_f[f"A{i}"] = ano-1
    ws_f[f"B{i}"] = fc
    ws_f[f"C{i}"] = fluxo_acumulado
    ws_f[f"D{i}"] = vp

    for col in ("B", "C", "D"):
        ws_f[f"{col}{i}"].number_format = 'R$ #,##0.00'

ws_f.column_dimensions["A"].width = 8
ws_f.column_dimensions["B"].width = 22
ws_f.column_dimensions["C"].width = 22
ws_f.column_dimensions["D"].width = 22

# ===== Aba INDICADORES =====
ws_i = wb.create_sheet("Indicadores_Pe_no_Chao")

ws_i["A1"] = "INDICADORES - CENÁRIO PÉ NO CHÃO"
ws_i["A1"].font = Font(bold=True, size=14)
ws_i.merge_cells("A1:C1")

ind_header = ["Indicador", "Descrição", "Resultado"]
for col, h in enumerate(ind_header, start=1):
    c = ws_i.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

lin = 3
ws_i[f"A{lin}"] = "Payback (anos)"
ws_i[f"A{lin}"].font = bold_font
ws_i[f"B{lin}"] = "Tempo para recuperar o investimento inicial pelo fluxo acumulado."
ws_i[f"C{lin}"] = float(payback_anos) if payback_anos is not None else None
ws_i[f"C{lin}"].number_format = '0,00'

lin += 1
ws_i[f"A{lin}"] = "VPL a 12%"
ws_i[f"A{lin}"].font = bold_font
ws_i[f"B{lin}"] = "Soma dos fluxos de caixa trazidos a valor presente."
ws_i[f"C{lin}"] = float(vpl_12)
ws_i[f"C{lin}"].number_format = 'R$ #,##0.00'

lin += 1
ws_i[f"A{lin}"] = "TIR"
ws_i[f"A{lin}"].font = bold_font
ws_i[f"B{lin}"] = "Taxa interna que zera o VPL do projeto."
ws_i[f"C{lin}"] = float(tir)
ws_i[f"C{lin}"].number_format = '0,00%'

lin += 1
ws_i[f"A{lin}"] = "Critério de aceitação"
ws_i[f"A{lin}"].font = bold_font
ws_i[f"B{lin}"] = "Projeto viável se VPL > 0 e TIR > taxa de desconto (12%)."
ws_i[f"C{lin}"] = "VIÁVEL" if (vpl_12 > 0 and tir > TAXA_DESCONTO_ANUAL) else "NÃO VIÁVEL"

ws_i.column_dimensions["A"].width = 26
ws_i.column_dimensions["B"].width = 60
ws_i.column_dimensions["C"].width = 18

# ---------------- SALVAR ARQUIVO ----------------
nome_arquivo = "Fluxo_Caixa_Lab_Eletronica_PE_NO_CHAO.xlsx"
wb.save(nome_arquivo)

print(f"\nArquivo gerado: {nome_arquivo}")
