import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ---------------- PARÂMETROS DO PROJETO ----------------
INVESTIMENTO_INICIAL = 53710     # R$
PRECO_HORA_ESTACAO = 35          # R$/h
HORAS_DIA = 4
DIAS_MES = 24
ESTACOES = 4

CUSTOS_FIXOS_MES = {
    'Aluguel': 1633,
    'Condomínio/IPTU': 1000,
    'Internet': 100,
    'Energia': 800,
    'Manutenção': 500,
    'Limpeza/Segurança': 300
}
CUSTOS_FIXOS_TOTAL = sum(CUSTOS_FIXOS_MES.values())

RECEITA_MES = PRECO_HORA_ESTACAO * HORAS_DIA * DIAS_MES * ESTACOES
LUCRO_OPERACIONAL_MES = RECEITA_MES - CUSTOS_FIXOS_TOTAL

TAXA_DESCONTO_ANUAL = 0.12
ANOS_ANALISE = 5

# ---------------- FUNÇÕES FINANCEIRAS ----------------
def calcular_vpl(fluxo, taxa):
    return sum(fc / (1 + taxa) ** t for t, fc in enumerate(fluxo))

def calcular_tir(fluxo):
    low, high = -0.99, 10.0
    for _ in range(100):
        mid = (low + high) / 2
        npv = calcular_vpl(fluxo, mid)
        if npv > 0:
            low = mid
        else:
            high = mid
    return low

def calcular_payback(fluxo):
    cumul = np.cumsum(fluxo)
    for i in range(1, len(cumul)):
        if cumul[i] >= 0:
            anos_completos = i - 1
            frac = abs(cumul[i-1]) / abs(fluxo[i])
            return anos_completos + frac
    return None  # não recupera

# ---------------- FLUXO DE CAIXA ANUAL ----------------
anos = list(range(0, ANOS_ANALISE + 1))
fluxo_caixa = [-INVESTIMENTO_INICIAL]
for ano in range(1, ANOS_ANALISE + 1):
    lucro_anual = LUCRO_OPERACIONAL_MES * 12 * (1 - 0.05 * (ano - 1))
    fluxo_caixa.append(lucro_anual)

payback_anos = calcular_payback(fluxo_caixa)
vpl_12 = calcular_vpl(fluxo_caixa, TAXA_DESCONTO_ANUAL)
tir = calcular_tir(fluxo_caixa)   # em fração (ex.: 0.45 = 45%)

# ---------------- CRIAÇÃO DA PLANILHA ----------------
wb = Workbook()

# ===== Aba PARÂMETROS =====
ws_p = wb.active
ws_p.title = "Parametros"

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
bold_font = Font(bold=True)

ws_p["A1"] = "PARÂMETROS DO PROJETO"
ws_p["A1"].font = Font(bold=True, size=14)
ws_p.merge_cells("A1:B1")

parametros = [
    ("Investimento Inicial (R$)", INVESTIMENTO_INICIAL),
    ("Preço/Hora/Estação (R$)", PRECO_HORA_ESTACAO),
    ("Horas por dia", HORAS_DIA),
    ("Dias por mês", DIAS_MES),
    ("Número de estações", ESTACOES),
    ("Receita mensal (R$)", RECEITA_MES),
    ("Custos fixos mensais (R$)", CUSTOS_FIXOS_TOTAL),
    ("Lucro operacional mensal (R$)", LUCRO_OPERACIONAL_MES),
    ("Taxa de desconto anual", TAXA_DESCONTO_ANUAL),
    ("Período de análise (anos)", ANOS_ANALISE),
]

for i, (nome, val) in enumerate(parametros, start=3):
    ws_p[f"A{i}"] = nome
    ws_p[f"A{i}"].font = bold_font
    ws_p[f"B{i}"] = float(val)
    if "R$" in nome:
        ws_p[f"B{i}"].number_format = 'R$ #,##0.00'
    elif "Taxa" in nome:
        ws_p[f"B{i}"].number_format = '0,00%'

# Cabeçalho visual
for col in ("A", "B"):
    cell = ws_p[f"{col}2"]
    cell.value = "Descrição" if col == "A" else "Valor"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

ws_p.column_dimensions["A"].width = 35
ws_p.column_dimensions["B"].width = 18

# ===== Aba FLUXO DE CAIXA =====
ws_f = wb.create_sheet("Fluxo de Caixa")

headers = ["Ano", "Fluxo de Caixa (R$)", "Fluxo Acumulado (R$)", "Valor Presente (R$)"]
for col, h in enumerate(headers, start=1):
    c = ws_f.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

fluxo_acumulado = 0
for i, ano in enumerate(anos, start=2):
    fc = fluxo_caixa[i-2]
    fluxo_acumulado += fc
    vp = fc / (1 + TAXA_DESCONTO_ANUAL) ** (ano)

    ws_f[f"A{i}"] = ano - 1
    ws_f[f"B{i}"] = fc
    ws_f[f"C{i}"] = fluxo_acumulado
    ws_f[f"D{i}"] = vp

    for col in ("B", "C", "D"):
        ws_f[f"{col}{i}"].number_format = 'R$ #,##0.00'

ws_f.column_dimensions["A"].width = 8
ws_f.column_dimensions["B"].width = 22
ws_f.column_dimensions["C"].width = 22
ws_f.column_dimensions["D"].width = 22

# ===== Aba INDICADORES =====
ws_i = wb.create_sheet("Indicadores")

ws_i["A1"] = "INDICADORES DE VIABILIDADE DO PROJETO"
ws_i["A1"].font = Font(bold=True, size=14)
ws_i.merge_cells("A1:C1")

indicadores_header = ["Indicador", "Descrição / Fórmula", "Resultado"]
for col, h in enumerate(indicadores_header, start=1):
    c = ws_i.cell(row=2, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

lin = 3
# Payback
ws_i[f"A{lin}"] = "Payback (anos)"
ws_i[f"A{lin}"].font = bold_font
ws_i[f"B{lin}"] = "Tempo para recuperar o investimento inicial pelo fluxo acumulado."
ws_i[f"C{lin}"] = float(payback_anos) if payback_anos is not None else None

lin += 1
# VPL
ws_i[f"A{lin}"] = "VPL a 12%"
ws_i[f"A{lin}"].font = bold_font
ws_i[f"B{lin}"] = "Soma dos fluxos de caixa trazidos a valor presente."
ws_i[f"C{lin}"] = float(vpl_12)
ws_i[f"C{lin}"].number_format = 'R$ #,##0.00'

lin += 1
# TIR
ws_i[f"A{lin}"] = "TIR"
ws_i[f"A{lin}"].font = bold_font
ws_i[f"B{lin}"] = "Taxa interna que zera o VPL do projeto."
ws_i[f"C{lin}"] = float(tir)
ws_i[f"C{lin}"].number_format = '0,00%'

lin += 1
# Critério
ws_i[f"A{lin}"] = "Critério de aceitação"
ws_i[f"A{lin}"].font = bold_font
ws_i[f"B{lin}"] = "Projeto viável se VPL>0 e TIR > taxa de desconto (12%)."
ws_i[f"C{lin}"] = "VIÁVEL" if (vpl_12 > 0 and tir > TAXA_DESCONTO_ANUAL) else "NÃO VIÁVEL"

ws_i.column_dimensions["A"].width = 22
ws_i.column_dimensions["B"].width = 55
ws_i.column_dimensions["C"].width = 18

# ---------------- SALVAR ARQUIVO ----------------
nome_arquivo = "Fluxo_Caixa_Lab_Eletronica_INDICADORES_OK.xlsx"
wb.save(nome_arquivo)

print(f"Arquivo gerado: {nome_arquivo}")
print(f"Payback (anos): {payback_anos:.2f}")
print(f"VPL (12%): R$ {vpl_12:,.2f}")
print(f"TIR: {tir*100:.2f}%")
